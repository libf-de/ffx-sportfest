# -*- coding: utf-8 -*-

"""
Module implementing ExcelImporter.
"""

from PyQt5.QtCore import pyqtSlot,  QRegExp
from PyQt5.QtGui import QRegExpValidator, QValidator
from PyQt5.QtWidgets import QDialog,  QProgressDialog,  QMessageBox,  QFileDialog
from openpyxl import load_workbook
from ui.TimerMessageBox import TimerMessageBox
from ui.KlasseInputDialog import KlasseInputDialog
from ui.GeschlechtInputDialog import GeschlechtInputDialog
from .Ui_ExcelDialog import Ui_Dialog
from lib.debug import Debug
import os, re,  json,  uuid, shutil,  string,  math,  codecs

wb = None
ws = None
tabv = None
gesv = None
klav = None
loadPD = None
procDbase = None

JSN = None

T1G = False
T2G = False
T3G = False
T4G = False

class ExcelImporter(QDialog, Ui_Dialog):
    """
    Class documentation goes here.
    """
    
    procDbase = None
    
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(ExcelImporter, self).__init__(parent)
        self.setupUi(self)
        
        #tblr = QRegExp("^(([A-Za-z]+)(\d+)\:([A-Za-z]+)(\d+))|(([A-Za-z]+)(\{\})\:([A-Za-z]+)(\{\}))$")
        #tblr = QRegExp("^(([A-Za-z]+)(\d+)\:([A-Za-z]+)(\d+))|(([A-Za-z]+)\*)$")
        tblr = QRegExp("^([A-Z]+)$")
        self.tabv = QRegExpValidator(tblr)
        gslr = QRegExp("^([Mm]|[Ww])$")
        self.gesv = QRegExpValidator(gslr)
        klsr = QRegExp("^([0-9]{1,2})\.([0-9]{1})$")
        self.klav = QRegExpValidator(klsr)
        self.nameEdit.setValidator(self.tabv)
        self.nameEdit.textChanged.connect(self.check_state_name)
        self.nameEdit.textChanged.emit(self.nameEdit.text())
        self.vornameEdit.setValidator(self.tabv)
        self.vornameEdit.textChanged.connect(self.check_state_vorname)
        self.vornameEdit.textChanged.emit(self.vornameEdit.text())
        self.geschlechtEdit.setValidator(self.tabv)
        self.geschlechtEdit.textChanged.connect(self.check_state_geschlecht)
        self.geschlechtEdit.textChanged.emit(self.geschlechtEdit.text())
        self.klasseEdit.setValidator(self.tabv)
        self.klasseEdit.textChanged.connect(self.check_state_klasse)
        self.klasseEdit.textChanged.emit(self.klasseEdit.text())
        
    def check_max(self, inp):
        if not self.ws is None:
            if self.getIndexFromColumn(inp) > self.ws.max_column:
                return False
            else:
                return True
        else:
            return True
        
    def check_state_name(self, *args, **kwargs):
        print("CS")
        sender = self.nameEdit
        validator = sender.validator()
        state = validator.validate(sender.text(), 0)[0]
        sval = False
        if state == QValidator.Acceptable and self.check_max(sender.text()):
            color = '#c4df9b' # green
            sval = True
        elif state == QValidator.Intermediate:
            color = '#fff79a' # yellow
        else:
            color = '#f6989d' # red
            
        self.T1G = sval
            
        self.considerLastStep()
        
        sender.setStyleSheet('QLineEdit { background-color: %s }' % color)
        
    def check_state_vorname(self, *args, **kwargs):
        print("CS")
        sender = self.vornameEdit
        validator = sender.validator()
        state = validator.validate(sender.text(), 0)[0]
        sval = False
        if state == QValidator.Acceptable and self.check_max(sender.text()):
            color = '#c4df9b' # green
            sval = True
        elif state == QValidator.Intermediate:
            color = '#fff79a' # yellow
        else:
            color = '#f6989d' # red
            
        self.T2G = sval
            
        self.considerLastStep()
        
        sender.setStyleSheet('QLineEdit { background-color: %s }' % color)
    
    def check_state_geschlecht(self, *args, **kwargs):
        print("CS")
        sender = self.geschlechtEdit
        validator = sender.validator()
        state = validator.validate(sender.text(), 0)[0]
        sval = False
        if state == QValidator.Acceptable and self.check_max(sender.text()):
            color = '#c4df9b' # green
            sval = True
        elif state == QValidator.Intermediate:
            color = '#fff79a' # yellow
        else:
            color = '#f6989d' # red
            
        self.T3G = sval
            
        self.considerLastStep()
        
        sender.setStyleSheet('QLineEdit { background-color: %s }' % color)
        
    def check_state_klasse(self, *args, **kwargs):
        print("CS")
        sender = self.klasseEdit
        validator = sender.validator()
        state = validator.validate(sender.text(), 0)[0]
        sval = False
        if state == QValidator.Acceptable and self.check_max(sender.text()):
            color = '#c4df9b' # green
            sval = True
        elif state == QValidator.Intermediate:
            color = '#fff79a' # yellow
        else:
            color = '#f6989d' # red
            
        self.T4G = sval
            
        self.considerLastStep()
        
        sender.setStyleSheet('QLineEdit { background-color: %s }' % color)
        
    def getSUID(self):
        return str(uuid.uuid4().hex[:8])
        
    def getIndexFromColumn(self,  colm):
        indx = 0
        for i in range(0, len(colm)):
            indx+=int((int(string.ascii_lowercase.index(colm[i].lower()))+1) * math.pow(26, i))
        return indx
        
    """
    Gibt die Spaltenbezeichnung (A,B,AA) für die angegebene Position zurück
    nach https://stackoverflow.com/questions/181596/how-to-convert-a-column-number-eg-127-into-an-excel-column-eg-aa
    @param ind Position
    """
    def getColumnFromIndex(self,  ind):
        divisor = int(ind)
        clm = ""
        while divisor > 0:
            modulo = (divisor - 1) % 26
            clm = str(chr(65 + modulo)) + clm
            divisor = int((divisor - modulo) / 26)
        return clm
    
    @pyqtSlot()
    def on_excelOpen_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        db_file,  _ = QFileDialog.getOpenFileName(self, "Datenbank öffnen", os.path.expanduser("~"), "Excel 2010+-Tabelle (*.xlsx);;Excel-Tabelle (*.xls);;Alle Dateien (*.*)");

        if db_file:
            self.excelPath.setText(str(db_file))
            self.parseExcel(db_file)
            
    def parseExcel(self,  file):
        self.loadPD = QProgressDialog()
        self.loadPD.setWindowTitle("Lade Tabelle...")
        self.loadPD.setLabelText("Lade Excel-Tabelle, bitte warten!")
        self.loadPD.setCancelButton(None)
        self.loadPD.setRange(0,  5)
        self.loadPD.show()
        self.loadPD.setValue(1)
        if not os.path.isfile(file):
            return False
        try:
            self.loadPD.setValue(2)
            self.wb = load_workbook(file)
            self.loadPD.setValue(3)
            self.sheetCombo.clear()
            self.sheetCombo.addItems(self.wb.get_sheet_names())
            self.loadPD.setValue(4)
            self.groupBox_4.setEnabled(True)
            self.sheetCombo.setEnabled(True)
            self.openSheet.setEnabled(True)
        except PermissionError as per:
            self.loadPD.cancel()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Die Tabelle konnte nicht geladen werden, da Sie keine Leserechte für diese Datei besitzen. Sie müssen Leserechte für diese Datei erlangen um sie zu öffnen.")
            msg.setWindowTitle("FFSportfest - Fehler beim Laden")
            msg.setDetailedText("Dateipfad: " + str(file) + "\n\n" + str(per))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        
        
        self.loadPD.setValue(5)
        self.loadPD.cancel()
    
    @pyqtSlot(bool)
    def on_geschlechtBox_toggled(self, checked):
        """
        Slot documentation goes here.
        
        @param checked DESCRIPTION
        @type bool
        """
        # TODO: not implemented yet
        if checked:
            self.geschlechtEdit.setValidator(self.tabv)
            self.geschlechtEdit.setText("")
            self.geschlechtEdit.setPlaceholderText("C")
        else:
            self.geschlechtEdit.setValidator(self.gesv)
            self.geschlechtEdit.setText("")
            self.geschlechtEdit.setPlaceholderText("M / W")
    
    @pyqtSlot(bool)
    def on_klasseBox_toggled(self, checked):
        """
        Slot documentation goes here.
        
        @param checked DESCRIPTION
        @type bool
        """
        # TODO: not implemented yet
        if checked:
            self.klasseEdit.setValidator(self.tabv)
            self.klasseEdit.setText("")
            self.klasseEdit.setPlaceholderText("D")
        else:
            self.klasseEdit.setValidator(self.klav)
            self.klasseEdit.setText("")
            self.klasseEdit.setPlaceholderText("5.3")
            
    def showCancelMessage(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("Der Vorgang wurde auf Ihren Wunsch hin abgebrochen!")
        msg.setWindowTitle("FFSportfest - Abbruch")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
        self.reject()
            
    def addExcelToJSON(self, dbPath):
        warns = []
        nme = []
        vne = []
        kls = []
        gsl = []
        
        nameRow = self.nameEdit.text()
        vnameRow = self.vornameEdit.text()
        
        if self.areaRows.isChecked():
            minrow = self.rowMin.value()
            maxrow = self.rowMax.value()
        else:
            minrow = self.ws.min_row
            maxrow = self.ws.max_row
            
        nameRange = str(nameRow + "{}:" + nameRow + "{}").format(minrow,maxrow)
        vornameRange = str(vnameRow + "{}:" + vnameRow + "{}").format(minrow,maxrow)
        geschlechtRange = str(self.geschlechtEdit.text() + "{}:" + self.geschlechtEdit.text() + "{}").format(minrow,maxrow)
        klasseRange = str(self.klasseEdit.text() + "{}:" + self.klasseEdit.text() + "{}").format(minrow,maxrow)
        
        for row in self.ws.iter_rows(nameRange):
            for cell in row:
                nme.append(str(cell.value))
                
        for row in self.ws.iter_rows(vornameRange):
            for cell in row:
                vne.append(str(cell.value))
                
        gslre = re.compile("^([Mm]|[Ww])$")
        GSLFB = None
        if self.geschlechtBox.isChecked():
            for row in self.ws.iter_rows(geschlechtRange):
                for cell in row:
                    if not gslre.match(str(cell.value)):
                        if GSLFB is None:
                            gd = GeschlechtInputDialog(self.ws[nameRow + cell.row].value, self.ws[vnameRow + cell.row].value,  cell.value)
                            if gd.exec_():
                                if gd.getForAll():
                                    GSLFB = gd.getSelection()
                                gsl.append(gd.getSelection())
                                warns.append("WARN: Ungültiges Geschlecht »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Korrektur auf " + gd.getSelection())
                            else:
                                self.showCancelMessage()
                                return
                        else:
                            warns.append("WARN: Ungültiges Geschlecht »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Nehme " + GSLFB + " an")
                            gsl.append(GSLFB)
                    else:
                        gsl.append(str(cell.value).upper())
        else:
            for i in range(minrow, maxrow):
                gsl.append(self.geschlechtEdit.text().upper())
        
        klsre = re.compile("^([0-9]{1,2})\.([0-9]{1})$")
        KLSFB = None
        if self.klasseBox.isChecked():
            for row in self.ws.iter_rows(klasseRange):
                for cell in row:
                    if not klsre.match(str(cell.value)):
                        if KLSFB is None:
                            kd = KlasseInputDialog(self.ws[nameRow + cell.row].value, self.ws[vnameRow + cell.row].value,  cell.value, cell.row)
                            if kd.exec_():
                                if kd.getReturnValue() == 1:
                                    self.showCancelMessage()
                                    return
                                elif kd.getReturnValue() == 2:
                                    kls.append(kd.getReturnKlasse())
                                    warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Korrigiert auf " + kd.getReturnKlasse())
                                elif kd.getReturnValue() == 3:
                                    kls.append(kd.getReturnKlasse())
                                    KLSFB = kd.getReturnKlasse()
                                    warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Standardkorrektur auf " + kd.getReturnKlasse() + " gesetzt")
                        else:
                            warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Nehme " + KLSFB + " an")
                            kls.append(KLSFB)
                    else:
                        kls.append(str(cell.value))
        else:
            for i in range(minrow, maxrow):
                kls.append(self.klasseEdit.text())
                
        if len(nme) == len(vne) == len(gsl) == len(kls):
            Debug.d("ExcelHinzufg: Gleiche Anzahl an Namen, Vornamen, Geschlecht und Klassen - fahre fort")
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Fehler: Die Anzahl der Namen, Vornamen, Geschlechten und Klassen stimmt nicht überein. Bitte stellen Sie sicher, das alle Tabellenbereiche gleich groß sind! Abbruch.")
            msg.setWindowTitle("FFSportfest - Fehler")
            msg.setDetailedText("Anzahl Namen: " + str(len(nme)) + "\nAnzahl Vornamen: " + str(len(vne)) + "\nAnzahl Geschlechte: " + str(len(gsl)) + "\nAnzahl Klassen: " + str(len(kls)))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
            
        for i in range(len(nme)):
            #uid name vorname geschlecht sprint_v/p/n lauf_v/p/n sprung_v/p/n wurf_v_p_n punkte note krank
            SUID = self.getSUID()
            SKLS = kls[i]
            
            if not SKLS in self.JSN:
                self.JSN[SKLS] = {}
            
            if not SUID in self.JSN[SKLS]:
                self.JSN[SKLS][SUID] = {}
            
            self.JSN[SKLS][SUID]['name'] = nme[i]
            self.JSN[SKLS][SUID]['vorname'] = vne[i]
            self.JSN[SKLS][SUID]['geschlecht'] = gsl[i]
            self.JSN[SKLS][SUID]['sprint_v'] = 0.0
            self.JSN[SKLS][SUID]['sprint_p'] = 0
            self.JSN[SKLS][SUID]['sprint_n'] = 6
            self.JSN[SKLS][SUID]['lauf_v'] = 0
            self.JSN[SKLS][SUID]['lauf_p'] = 0
            self.JSN[SKLS][SUID]['lauf_n'] = 6
            self.JSN[SKLS][SUID]['sprung_v'] = 0.0
            self.JSN[SKLS][SUID]['sprung_p'] = 0
            self.JSN[SKLS][SUID]['sprung_n'] = 6
            self.JSN[SKLS][SUID]['wurf_v'] = 0.0
            self.JSN[SKLS][SUID]['wurf_p'] = 0
            self.JSN[SKLS][SUID]['wurf_n'] = 6
            self.JSN[SKLS][SUID]['punkte'] = 0
            self.JSN[SKLS][SUID]['note'] = 6
            self.JSN[SKLS][SUID]['krank'] = False
            
        if len(warns) > 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("FFSportfest - Aus Excel importieren")
            msg.setText("Beim Verarbeiten der Excel-Tabelle sind " + str(len(warns)) + " Warnungen aufgetreten (siehe Details). Möchten Sie den Import fortsetzen und die veränderte Datenbank abspeichern? Es ist möglicherweise nötig die Datenbank im Nachhinein zu korrigieren.")
            msg.setDetailedText("\n".join(warns))
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            rt = msg.exec_()
            
            if rt == QMessageBox.No:
                return
        try:        
            with open(dbPath, 'w', encoding='utf8') as f:
                json.dump(self.JSN, f, indent=4, sort_keys=True, ensure_ascii=False)
                
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("FFSportfest - Fertig")
            msg.setText("Die Daten wurden erfolgreich aus der Excel-Tabelle in die Datenbank " + str(os.path.basename(dbPath)) + " importiert. Es wurden " + str(len(nme)) + " neue Schüler hinzugefügt")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            self.procDbase = dbPath
            self.accept()
        except PermissionError as per:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Die Datenbank konnte nicht gespeichert werden, da Sie keine Schreibrechte für diesen Ort besitzen. Sie müssen Schreibrechte für diesen Ort erlangen um Daten in diese Datenbank zu importieren.")
            msg.setWindowTitle("FFSportfest - Fehler beim Schreiben")
            msg.setDetailedText("Dateipfad: " + str(dbPath) + "\n\n" + str(per))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        except Exception as exc:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Die Datenbank konnte nicht gespeichert werden.")
            msg.setWindowTitle("FFSportfest - Fehler beim Speichern")
            msg.setDetailedText(str(exc))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            
    def getReturnDb(self):
        return self.procDbase
    
    def newExcelJSON(self):
        warns = []
        nme = []
        vne = []
        kls = []
        gsl = []
        
        nameRow = self.nameEdit.text()
        vnameRow = self.vornameEdit.text()
        
        if self.areaRows.isChecked():
            minrow = self.rowMin.value()
            maxrow = self.rowMax.value()
        else:
            minrow = self.ws.min_row
            maxrow = self.ws.max_row
            
        nameRange = str(nameRow + "{}:" + nameRow + "{}").format(minrow,maxrow)
        vornameRange = str(vnameRow + "{}:" + vnameRow + "{}").format(minrow,maxrow)
        geschlechtRange = str(self.geschlechtEdit.text() + "{}:" + self.geschlechtEdit.text() + "{}").format(minrow,maxrow)
        klasseRange = str(self.klasseEdit.text() + "{}:" + self.klasseEdit.text() + "{}").format(minrow,maxrow)
        
        for row in self.ws.iter_rows(nameRange):
            for cell in row:
                nme.append(str(cell.value))
                
        for row in self.ws.iter_rows(vornameRange):
            for cell in row:
                vne.append(str(cell.value))
                
        gslre = re.compile("^([Mm]|[Ww])$")
        GSLFB = None
        if self.geschlechtBox.isChecked():
            for row in self.ws.iter_rows(geschlechtRange):
                for cell in row:
                    if not gslre.match(str(cell.value)):
                        if GSLFB is None:
                            gd = GeschlechtInputDialog(self.ws[nameRow + str(cell.row)].value, self.ws[vnameRow + str(cell.row)].value,  cell.value)
                            if gd.exec_():
                                if gd.getForAll():
                                    GSLFB = gd.getSelection()
                                gsl.append(gd.getSelection())
                                warns.append("WARN: Ungültiges Geschlecht »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Korrektur auf " + gd.getSelection())
                            else:
                                self.showCancelMessage()
                                return
                        else:
                            warns.append("WARN: Ungültiges Geschlecht »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Nehme " + GSLFB + " an")
                            gsl.append(GSLFB)
                    else:
                        gsl.append(str(cell.value).upper())
        else:
            for i in range(minrow, maxrow):
                gsl.append(self.geschlechtEdit.text().upper())
        
        klsre = re.compile("^([0-9]{1,2})\.([0-9]{1})$")
        KLSFB = None
        if self.klasseBox.isChecked():
            for row in self.ws.iter_rows(klasseRange):
                for cell in row:
                    if not klsre.match(str(cell.value)):
                        if KLSFB is None:
                            kd = KlasseInputDialog(self.ws[nameRow + str(cell.row)].value, self.ws[vnameRow + str(cell.row)].value,  cell.value, cell.row)
                            if kd.exec_():
                                if kd.getReturnValue() == 1:
                                    self.showCancelMessage()
                                    return
                                elif kd.getReturnValue() == 2:
                                    kls.append(kd.getReturnKlasse())
                                    warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Korrigiert auf " + kd.getReturnKlasse())
                                elif kd.getReturnValue() == 3:
                                    kls.append(kd.getReturnKlasse())
                                    KLSFB = kd.getReturnKlasse()
                                    warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Standardkorrektur auf " + kd.getReturnKlasse() + " gesetzt")
                        else:
                            warns.append("WARN: Ungültige Klasse »" + str(cell.value) + "« auf Zeile " + str(cell.row) + ": Nehme " + KLSFB + " an")
                            kls.append(KLSFB)
                    else:
                        kls.append(str(cell.value))
        else:
            for i in range(minrow, maxrow):
                kls.append(self.klasseEdit.text())
                
        if len(nme) == len(vne) == len(gsl) == len(kls):
            Debug.d("ExcelHinzufg: Gleiche Anzahl an Namen, Vornamen, Geschlecht und Klassen - fahre fort")
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Fehler: Die Anzahl der Namen, Vornamen, Geschlechten und Klassen stimmt nicht überein. Bitte stellen Sie sicher, das alle Tabellenbereiche gleich groß sind! Abbruch.")
            msg.setWindowTitle("FFSportfest - Fehler")
            msg.setDetailedText("Anzahl Namen: " + str(len(nme)) + "\nAnzahl Vornamen: " + str(len(vne)) + "\nAnzahl Geschlechte: " + str(len(gsl)) + "\nAnzahl Klassen: " + str(len(kls)))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
            
        self.JSN = {}
        
        for i in range(len(nme)):
            #uid name vorname geschlecht sprint_v/p/n lauf_v/p/n sprung_v/p/n wurf_v_p_n punkte note krank
            SUID = self.getSUID()
            SKLS = kls[i]
            
            #SUID_DIC = { 'name' : nme[i], 'vorname' : vne[i], 'geschlecht' : gsl[i], }
            
            if not SKLS in self.JSN:
                self.JSN[SKLS] = {}
            
            if not SUID in self.JSN[SKLS]:
                self.JSN[SKLS][SUID] = {}
            
            self.JSN[SKLS][SUID]['name'] = nme[i]
            self.JSN[SKLS][SUID]['vorname'] = vne[i]
            self.JSN[SKLS][SUID]['geschlecht'] = gsl[i]
            self.JSN[SKLS][SUID]['sprint_v'] = 0.0
            self.JSN[SKLS][SUID]['sprint_p'] = 0
            self.JSN[SKLS][SUID]['sprint_n'] = 6
            self.JSN[SKLS][SUID]['lauf_v'] = 0
            self.JSN[SKLS][SUID]['lauf_p'] = 0
            self.JSN[SKLS][SUID]['lauf_n'] = 6
            self.JSN[SKLS][SUID]['sprung_v'] = 0.0
            self.JSN[SKLS][SUID]['sprung_p'] = 0
            self.JSN[SKLS][SUID]['sprung_n'] = 6
            self.JSN[SKLS][SUID]['wurf_v'] = 0.0
            self.JSN[SKLS][SUID]['wurf_p'] = 0
            self.JSN[SKLS][SUID]['wurf_n'] = 6
            self.JSN[SKLS][SUID]['punkte'] = 0
            self.JSN[SKLS][SUID]['note'] = 6
            self.JSN[SKLS][SUID]['krank'] = False
            
        if len(warns) > 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("FFSportfest - Neue Datenbank aus Excel")
            msg.setText("Beim Verarbeiten der Excel-Tabelle sind " + str(len(warns)) + " Warnungen aufgetreten (siehe Details). Es ist möglicherweise nötig die Datenbank im Nachhinein zu korrigieren. Datenbank erstellen?")
            msg.setDetailedText("\n".join(warns))
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            rt = msg.exec_()
            
            if rt == QMessageBox.No:
                return
                
        nDbFile,  _ = QFileDialog.getSaveFileName(self, "FFSportfest - Neue Datenbank aus Excel speichern", os.path.expanduser("~"), "FFD-Datenbank (*.ffd)");

        if nDbFile:
            try:
                with open(nDbFile, 'w',  encoding='utf8') as f:
                        json.dump(self.JSN, f, indent=4, sort_keys=True, ensure_ascii=False)                
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Question)
                msg.setWindowTitle("FFSportfest - Fertig")
                msg.setText("Die Datenbank " + str(os.path.basename(nDbFile)) + " wurde erfolgreich aus der Excel-Tabelle erstellt. Es wurden " + str(len(nme)) + " Schüler hinzugefügt")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                self.procDbase = nDbFile
                self.accept()
            except PermissionError as per:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Die Datenbank konnte nicht gespeichert werden, da Sie keine Schreibrechte für diesen Ort besitzen. Sie müssen Schreibrechte für diesen Ort erlangen oder die Datenbank an einem anderen Ort speichern.")
                msg.setWindowTitle("FFSportfest - Fehler beim Schreiben")
                msg.setDetailedText("Dateipfad: " + str(nDbFile) + "\n\n" + str(per))
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
            """except Exception as exc:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Die Datenbank konnte nicht gespeichert werden.")
                msg.setWindowTitle("FFSportfest - Fehler beim Speichern")
                msg.setDetailedText(str(exc))
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()"""
        else:
            self.showCancelMessage()
        
    def loadDb(self,  path):
        self.loadPD = QProgressDialog()
        self.loadPD.setWindowTitle("Lade Datenbank...")
        self.loadPD.setLabelText("Lade Datenbank, bitte warten!")
        self.loadPD.setCancelButton(None)
        self.loadPD.setRange(0,  5)
        self.loadPD.show()
        self.loadPD.setValue(1)
        if not os.path.isfile(path):
            return False
        try:
            self.loadPD.setValue(2)
            self.JSN = json.load(codecs.open(str(path), 'r', 'utf-8-sig'))
            #self.JSN = json.load(open(str(path)))
            self.loadPD.setValue(3)
        except PermissionError as per:
            self.loadPD.cancel()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Die Datenbank konnte nicht geladen werden, da Sie keine Leserechte für diese Datei besitzen. Sie müssen Leserechte für diese Datei erlangen um sie zu öffnen.")
            msg.setWindowTitle("FFSportfest - Fehler beim Laden")
            msg.setDetailedText("Dateipfad: " + str(self.dbPath) + "\n\n" + str(per))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        except json.decoder.JSONDecodeError as exc:
            self.loadPD.cancel()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Die Datenbank konnte nicht geladen werden! Sie ist möglicherweise beschädigt.")
            msg.setInformativeText("Sie können versuchen, falls aktiviert und vorhanden, die zugehörige Sicherungsdatei (Dateiname.ffd~) zu öffnen oder die Datenbank per Hand mit einem Texteditor zu reparieren. Klicken Sie auf Details anzeigen um zu sehen wo und welcher Fehler aufgetreten ist.")
            msg.setWindowTitle("FFSportfest - Fehler beim Öffnen")
            msg.setDetailedText(str(exc))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        try:
                shutil.copy2(str(path), str(path) + "~") 
        except PermissionError as perb:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Es konnte keine Sicherungsdatei erstellt werden, da Sie nicht die nötigen Berechtigungen für den Ordner besitzen. Die Datenbank sollte aber trotzdem problemlos geladen worden sein.")
                msg.setWindowTitle("FFSportfest - Warnung")
                msg.setDetailedText("Quellpfad: " + str(path) + "\n\n" + str(perb))
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
        self.loadPD.cancel()
        self.addExcelToJSON(path)
    
    @pyqtSlot()
    def on_createNew_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        self.newExcelJSON()
    
    @pyqtSlot()
    def on_addInto_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        db_file,  _ = QFileDialog.getOpenFileName(self, "Datenbank öffnen", os.path.expanduser("~"), "FFD-Datenbank (*.ffd);;Alle Dateien (*.*)");

        if db_file:
            self.loadDb(db_file)
            
            if self.JSN is None:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Existierende Datenbank konnte nicht geladen werden - Abbruch!")
                msg.setWindowTitle("FFSportfest")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return
            
            
    
    @pyqtSlot()
    def on_openSheet_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        try:
            self.ws = self.wb[self.sheetCombo.currentText()]
            self.groupBox_2.setEnabled(True)
            self.label_2.setEnabled(True)
            self.label_4.setEnabled(True)
            self.label_3.setEnabled(True)
            self.vornameEdit.setEnabled(True)
            self.nameEdit.setEnabled(True)
            self.geschlechtBox.setEnabled(True)
            self.geschlechtEdit.setEnabled(True)
            self.label_5.setEnabled(True)
            self.klasseBox.setEnabled(True)
            self.klasseEdit.setEnabled(True)
            self.groupBox_5.setEnabled(True)
            self.allRows.setEnabled(True)
            self.areaRows.setEnabled(True)
            self.rowMin.setMaximum(self.ws.max_row)
            self.rowMax.setMaximum(self.ws.max_row)
            print("RC: {}, CC: {}".format(self.ws.max_row, self.ws.max_column))
        except Exception as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Das Arbeitsblatt konnte nicht geladen werden.")
            msg.setWindowTitle("FFSportfest - Fehler beim Laden")
            msg.setDetailedText(str(e))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        
    def considerLastStep(self):
        if self.T1G and self.T2G and self.T3G and self.T4G:
            self.groupBox_3.setEnabled(True)
            self.createNew.setEnabled(True)
            self.addInto.setEnabled(True)
            self.groupBox_6.setEnabled(True)
            self.vornameLabel.setEnabled(True)
            self.vornamePreview.setEnabled(True)
            self.nachnameLabel.setEnabled(True)
            self.nachnamePreview.setEnabled(True)
            self.geschlechtLabel.setEnabled(True)
            self.geschlechtPreview.setEnabled(True)
            self.klasseLabel.setEnabled(True)
            self.klassePreview.setEnabled(True)
            self.updatePreview()
        else:
            self.groupBox_3.setEnabled(False)
            self.createNew.setEnabled(False)
            self.addInto.setEnabled(False)
            self.groupBox_6.setEnabled(False)
            self.vornameLabel.setEnabled(False)
            self.vornamePreview.setEnabled(False)
            self.nachnameLabel.setEnabled(False)
            self.nachnamePreview.setEnabled(False)
            self.geschlechtLabel.setEnabled(False)
            self.geschlechtPreview.setEnabled(False)
            self.klasseLabel.setEnabled(False)
            self.klassePreview.setEnabled(False)
    
    def updatePreview(self):
        firstRow = "1"
        if self.areaRows.isChecked():
            firstRow = str(self.rowMin.value())
        
        nameCol = self.nameEdit.text()
        vornameCol = self.vornameEdit.text()
        gsl = self.geschlechtEdit.text()
        kls = self.klasseEdit.text()
    
        self.nachnamePreview.setText(str(self.ws[nameCol + firstRow].value))
        self.vornamePreview.setText(str(self.ws[vornameCol + firstRow].value))
        if self.geschlechtBox.isChecked():
            self.geschlechtPreview.setText(str(self.ws[gsl + firstRow].value))
        else:
            self.geschlechtPreview.setText(gsl)
        
        if self.klasseBox.isChecked():
            self.klassePreview.setText(str(self.ws[kls + firstRow].value))
        else:
            self.klassePreview(kls)
        
        
    
    @pyqtSlot(int)
    def on_rowMin_valueChanged(self, p0):
        """
        Slot documentation goes here.
        
        @param p0 DESCRIPTION
        @type int
        """
        # TODO: not implemented yet
        self.rowMax.setMinimum(p0)
        if self.T1G and self.T2G and self.T3G and self.T4G:
            self.updatePreview()
    
    @pyqtSlot(int)
    def on_rowMax_valueChanged(self, p0):
        """
        Slot documentation goes here.
        
        @param p0 DESCRIPTION
        @type int
        """
        if self.T1G and self.T2G and self.T3G and self.T4G:
            self.updatePreview()
        
    
    @pyqtSlot()
    def on_vornameEdit_editingFinished(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        if not self.ws is None:
            if self.getIndexFromColumn(self.vornameEdit.text()) > self.ws.max_column:
                self.vornameEdit.setFocus()
                self.vornameEdit.setText(self.getColumnFromIndex(self.ws.max_column))
                mb = TimerMessageBox(2, self)
                mb.setText("Eingebene Spalte existiert nicht - Wert wurde auf letzte Spalte gesetzt!")
                mb.setIcon(QMessageBox.Warning)
                mb.exec_()
                
            
    
    @pyqtSlot()
    def on_nameEdit_editingFinished(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        if not self.ws is None:
            if self.getIndexFromColumn(self.nameEdit.text()) > self.ws.max_column:
                self.nameEdit.setFocus()
                self.nameEdit.setText(self.getColumnFromIndex(self.ws.max_column))
                mb = TimerMessageBox(2, self)
                mb.setText("Eingebene Spalte existiert nicht - Wert wurde auf letzte Spalte gesetzt!")
                mb.setIcon(QMessageBox.Warning)
                mb.exec_()
    
    @pyqtSlot() 
    def on_geschlechtEdit_editingFinished(self):
        """
        Slot documentation goes here.
        """
        if not self.ws is None:
            if self.geschlechtBox.isChecked():
                if self.getIndexFromColumn(self.geschlechtEdit.text()) > self.ws.max_column:
                    self.geschlechtEdit.setFocus()
                    self.geschlechtEdit.setText(self.getColumnFromIndex(self.ws.max_column))
                    mb = TimerMessageBox(2, self)
                    mb.setText("Eingebene Spalte existiert nicht - Wert wurde auf letzte Spalte gesetzt!")
                    mb.setIcon(QMessageBox.Warning)
                    mb.exec_()
    
    @pyqtSlot()
    def on_klasseEdit_editingFinished(self):
        """
        Slot documentation goes here.
        """
        if not self.ws is None:
            if self.klasseBox.isChecked():
                if self.getIndexFromColumn(self.klasseEdit.text()) > self.ws.max_column:
                    self.klasseEdit.setFocus()
                    self.klasseEdit.setText(self.getColumnFromIndex(self.ws.max_column))
                    mb = TimerMessageBox(2, self)
                    mb.setText("Eingebene Spalte existiert nicht - Wert wurde auf letzte Spalte gesetzt!")
                    mb.setIcon(QMessageBox.Warning)
                    mb.exec_()
    
    @pyqtSlot(bool)
    def on_areaRows_toggled(self, checked):
        """
        Slot documentation goes here.
        
        @param checked DESCRIPTION
        @type bool
        """
        self.rowMin.setEnabled(checked)
        self.rowMax.setEnabled(checked)
        if self.T1G and self.T2G and self.T3G and self.T4G:
            self.updatePreview()
